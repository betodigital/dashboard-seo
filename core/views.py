import csv
import io
import json
import urllib.request
import urllib.error
from datetime import datetime, timedelta, timezone

import pytz

from django.conf import settings as django_settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from django.utils import timezone as dj_timezone
from functools import wraps

from .models import UserProfile, AccessLog
from .config_utils import get_config, get_all_configs, set_config
from .email_utils import send_test_email

SP_TZ = pytz.timezone('America/Sao_Paulo')


# ── Decorador de permissão por tela ──────────────────────────────
def require_permission(screen):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            try:
                if not request.user.profile.has_access(screen):
                    messages.error(request, 'Voce nao tem permissao para acessar esta tela.')
                    return redirect('dashboard')
            except UserProfile.DoesNotExist:
                messages.error(request, 'Perfil nao encontrado.')
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        return _wrapped
    return decorator


# ── Root redirect ─────────────────────────────────────────────────
def redirect_root(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


# ── PWA — manifest.json dinâmico ──────────────────────────────────
def pwa_manifest(request):
    system_name = get_config('system_name', 'MyApp') or 'MyApp'
    manifest = {
        "name": system_name,
        "short_name": system_name[:12],
        "description": f"Aplicativo {system_name}",
        "start_url": "/dashboard/",
        "display": "standalone",
        "background_color": "#0a0a0f",
        "theme_color": "#6c63ff",
        "orientation": "any",
        "icons": [
            {"src": "/static/icons/icon-72x72.png",   "sizes": "72x72",   "type": "image/png"},
            {"src": "/static/icons/icon-96x96.png",   "sizes": "96x96",   "type": "image/png"},
            {"src": "/static/icons/icon-128x128.png", "sizes": "128x128", "type": "image/png"},
            {"src": "/static/icons/icon-144x144.png", "sizes": "144x144", "type": "image/png"},
            {"src": "/static/icons/icon-152x152.png", "sizes": "152x152", "type": "image/png"},
            {"src": "/static/icons/icon-192x192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/static/icons/icon-384x384.png", "sizes": "384x384", "type": "image/png"},
            {"src": "/static/icons/icon-512x512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
        "screenshots": [],
        "categories": ["productivity", "business"],
    }
    return JsonResponse(manifest, content_type='application/manifest+json')


# ── PWA — Service Worker (deve ser servido da raiz) ───────────────
def pwa_service_worker(request):
    import os
    sw_path = os.path.join(django_settings.BASE_DIR, 'static', 'js', 'sw.js')
    with open(sw_path, 'r', encoding='utf-8') as f:
        content = f.read()
    response = HttpResponse(content, content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    return response


# ── Login ─────────────────────────────────────────────────────────
@never_cache
@require_http_methods(['GET', 'POST'])
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        if not username or not password:
            messages.error(request, 'Preencha usuario e senha.')
            return render(request, 'login.html')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            if not user.is_active:
                messages.error(request, 'Conta desativada. Contate o administrador.')
                return render(request, 'login.html')
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Usuario ou senha incorretos.')

    return render(request, 'login.html')


# ── Dashboard ─────────────────────────────────────────────────────
@login_required
@never_cache
def dashboard_view(request):
    return render(request, 'dashboard.html', {
        'last_login': request.user.last_login,
    })


# ── Relatorios ────────────────────────────────────────────────────
@require_permission('reports')
@never_cache
def reports_view(request):
    return render(request, 'reports.html')


# ── Configuracoes ─────────────────────────────────────────────────
SMTP_KEYS    = ['smtp_host','smtp_port','smtp_user','smtp_password','smtp_use_tls','smtp_use_ssl','email_from','email_from_name']
DB_CFG_KEYS  = ['db_host','db_port','db_name','db_user','db_password']
AI_KEYS      = ['openai_api_key']
GENERAL_KEYS = ['system_name', 'app_url']


def _settings_base_ctx():
    """Contexto base compartilhado por todas as views que renderizam settings.html."""
    smtp_cfg    = get_all_configs(SMTP_KEYS)
    db_cfg      = get_all_configs(DB_CFG_KEYS)
    ai_cfg      = get_all_configs(AI_KEYS)
    general_cfg = get_all_configs(GENERAL_KEYS)
    db_settings = django_settings.DATABASES.get('default', {})
    return {
        'smtp':          smtp_cfg,
        'db_cfg':        db_cfg,
        'ai_cfg':        ai_cfg,
        'general_cfg':   general_cfg,
        'has_openai_key': bool(ai_cfg.get('openai_api_key', '').strip()),
        'db_current': {
            'host':   db_settings.get('HOST', 'localhost'),
            'port':   db_settings.get('PORT', '5432'),
            'name':   db_settings.get('NAME', ''),
            'user':   db_settings.get('USER', ''),
            'engine': db_settings.get('ENGINE', '').split('.')[-1],
        },
        # logs tab defaults (evita VariableDoesNotExist nas outras abas)
        'logs':          [],
        'stats':         {'total': 0, 'login_ok': 0, 'login_fail': 0, 'logout': 0, 'lockout': 0},
        'date_from_str': '',
        'date_to_str':   '',
        'event_filter':  '',
        'user_filter':   '',
        'period':        '',
        'now_sp':        '',
        'event_choices': AccessLog.EVENT_CHOICES,
    }


@require_permission('settings')
@never_cache
@require_http_methods(['GET', 'POST'])
def settings_view(request):
    active_tab = request.GET.get('tab', 'general')

    if request.method == 'POST':
        form_type = request.POST.get('form_type', '')

        if form_type == 'general':
            set_config('system_name', request.POST.get('system_name', '').strip())
            set_config('app_url', request.POST.get('app_url', '').strip())
            messages.success(request, 'Configuracoes gerais salvas com sucesso!')
            return redirect('/settings/?tab=general')

        elif form_type == 'email':
            for key in SMTP_KEYS:
                set_config(key, request.POST.get(key, ''))
            # TLS/SSL são checkboxes — ausente = '0'
            set_config('smtp_use_tls', '1' if 'smtp_use_tls' in request.POST else '0')
            set_config('smtp_use_ssl', '1' if 'smtp_use_ssl' in request.POST else '0')
            messages.success(request, 'Configuracoes de e-mail salvas com sucesso!')
            return redirect(f'/settings/?tab=email')

        elif form_type == 'database':
            for key in DB_CFG_KEYS:
                set_config(key, request.POST.get(key, ''))
            messages.success(request, 'Configuracoes de banco salvas. Reinicie o servidor para aplicar.')
            return redirect(f'/settings/?tab=database')

        elif form_type == 'openai':
            set_config('openai_api_key', request.POST.get('openai_api_key', ''))
            messages.success(request, 'Configuracoes de IA salvas com sucesso!')
            return redirect(f'/settings/?tab=openai')

    # GET — carrega valores salvos
    ctx = _settings_base_ctx()
    ctx['active_tab'] = active_tab
    return render(request, 'settings.html', ctx)


# ── Configuracoes — Teste de E-mail (AJAX) ────────────────────────
@require_permission('settings')
@require_http_methods(['POST'])
def test_email_view(request):
    to_email = request.POST.get('to_email', '').strip() or request.user.email
    if not to_email:
        return JsonResponse({'ok': False, 'error': 'Nenhum e-mail de destino informado.'})
    ok, error = send_test_email(to_email)
    return JsonResponse({'ok': ok, 'error': error or ''})


# ── Configuracoes — Buscar Modelos OpenAI (AJAX) ──────────────────
@require_permission('settings')
@require_http_methods(['POST'])
def fetch_openai_models_view(request):
    api_key = get_config('openai_api_key', '').strip()
    if not api_key:
        return JsonResponse({'ok': False, 'error': 'Chave de API nao configurada.'})

    url = 'https://api.openai.com/v1/models'
    req = urllib.request.Request(url, headers={
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
        all_models = [m['id'] for m in data.get('data', [])]
        # Filtra somente modelos de linguagem relevantes
        prefixes = ('gpt-', 'o1', 'o3', 'o4', 'chatgpt')
        filtered = sorted([m for m in all_models if any(m.startswith(p) for p in prefixes)])
        return JsonResponse({'ok': True, 'models': filtered})
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        try:
            msg = json.loads(body).get('error', {}).get('message', str(e))
        except Exception:
            msg = str(e)
        return JsonResponse({'ok': False, 'error': f'Erro da API OpenAI: {msg}'})
    except urllib.error.URLError as e:
        return JsonResponse({'ok': False, 'error': f'Erro de conexao: {e.reason}'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# ── Usuarios — Lista ──────────────────────────────────────────────
@require_permission('users')
@never_cache
def users_list_view(request):
    users = User.objects.select_related('profile').order_by('-date_joined')
    return render(request, 'users/list.html', {'users': users})


# ── Usuarios — Criar ──────────────────────────────────────────────
@require_permission('users')
@never_cache
@require_http_methods(['GET', 'POST'])
def user_create_view(request):
    if request.method == 'POST':
        username  = request.POST.get('username', '').strip()
        email     = request.POST.get('email', '').strip()
        full_name = request.POST.get('full_name', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        errors = []
        if not username:
            errors.append('Usuario e obrigatorio.')
        if User.objects.filter(username=username).exists():
            errors.append('Este nome de usuario ja existe.')
        if not email:
            errors.append('E-mail e obrigatorio.')
        if User.objects.filter(email=email).exists():
            errors.append('Este e-mail ja esta em uso.')
        if len(password1) < 8:
            errors.append('A senha deve ter ao menos 8 caracteres.')
        if password1 != password2:
            errors.append('As senhas nao coincidem.')

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'users/form.html', {'form_data': request.POST, 'is_new': True})

        user = User.objects.create_user(username=username, email=email, password=password1)
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.full_name          = full_name
        profile.can_view_dashboard = 'can_view_dashboard' in request.POST
        profile.can_view_reports   = 'can_view_reports'   in request.POST
        profile.can_view_settings  = 'can_view_settings'  in request.POST
        profile.can_view_users     = 'can_view_users'     in request.POST
        profile.save()

        messages.success(request, f'Usuario "{username}" criado com sucesso!')
        return redirect('users_list')

    return render(request, 'users/form.html', {'form_data': {}, 'is_new': True})


# ── Usuarios — Editar ─────────────────────────────────────────────
@require_permission('users')
@never_cache
@require_http_methods(['GET', 'POST'])
def user_edit_view(request, user_id):
    target  = get_object_or_404(User, pk=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=target)

    if request.method == 'POST':
        email     = request.POST.get('email', '').strip()
        full_name = request.POST.get('full_name', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        is_active = 'is_active' in request.POST

        errors = []
        if not email:
            errors.append('E-mail e obrigatorio.')
        if User.objects.filter(email=email).exclude(pk=target.pk).exists():
            errors.append('Este e-mail ja esta em uso.')
        if password1 and len(password1) < 8:
            errors.append('A senha deve ter ao menos 8 caracteres.')
        if password1 and password1 != password2:
            errors.append('As senhas nao coincidem.')

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'users/form.html', {
                'target': target, 'profile': profile,
                'form_data': request.POST, 'is_new': False
            })

        target.email     = email
        target.is_active = is_active
        if password1:
            target.set_password(password1)
        target.save()

        profile.full_name          = full_name
        profile.can_view_dashboard = 'can_view_dashboard' in request.POST
        profile.can_view_reports   = 'can_view_reports'   in request.POST
        profile.can_view_settings  = 'can_view_settings'  in request.POST
        profile.can_view_users     = 'can_view_users'     in request.POST
        profile.save()

        messages.success(request, f'Usuario "{target.username}" atualizado!')
        return redirect('users_list')

    return render(request, 'users/form.html', {
        'target': target, 'profile': profile,
        'form_data': {}, 'is_new': False
    })


# ── Usuarios — Excluir ────────────────────────────────────────────
@require_permission('users')
@require_http_methods(['POST'])
def user_delete_view(request, user_id):
    target = get_object_or_404(User, pk=user_id)
    if target == request.user:
        messages.error(request, 'Voce nao pode excluir sua propria conta.')
        return redirect('users_list')
    if target.is_superuser:
        messages.error(request, 'Nao e possivel excluir um superusuario.')
        return redirect('users_list')
    username = target.username
    target.delete()
    messages.success(request, f'Usuario "{username}" excluido.')
    return redirect('users_list')


# ── Log de Acesso ─────────────────────────────────────────────────
def _parse_log_filters(request):
    """Retorna (date_from, date_to) como datetimes UTC a partir dos parâmetros GET."""
    now_sp = dj_timezone.now().astimezone(SP_TZ)

    # Período predefinido
    period = request.GET.get('period', '')
    if period == '7d':
        date_from = now_sp - timedelta(days=7)
        date_to   = now_sp
    elif period == '30d':
        date_from = now_sp - timedelta(days=30)
        date_to   = now_sp
    elif period == '90d':
        date_from = now_sp - timedelta(days=90)
        date_to   = now_sp
    else:
        # Datas manuais
        from_str = request.GET.get('date_from', '')
        to_str   = request.GET.get('date_to', '')
        try:
            date_from = SP_TZ.localize(datetime.strptime(from_str, '%Y-%m-%d'))
        except (ValueError, TypeError):
            date_from = now_sp - timedelta(days=30)
        try:
            date_to = SP_TZ.localize(datetime.strptime(to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except (ValueError, TypeError):
            date_to = now_sp

    return date_from.astimezone(pytz.utc), date_to.astimezone(pytz.utc)


def _format_timestamp(ts):
    """Converte timestamp UTC para horário de São Paulo formatado."""
    if ts is None:
        return ''
    if ts.tzinfo is None:
        ts = pytz.utc.localize(ts)
    return ts.astimezone(SP_TZ).strftime('%d/%m/%Y %H:%M:%S')


@require_permission('settings')
@never_cache
def logs_view(request):
    date_from, date_to = _parse_log_filters(request)

    qs = AccessLog.objects.filter(
        timestamp__gte=date_from,
        timestamp__lte=date_to,
    ).select_related('user').order_by('-timestamp')

    # Filtro de evento
    event_filter = request.GET.get('event', '')
    if event_filter:
        qs = qs.filter(event=event_filter)

    # Filtro de usuário
    user_filter = request.GET.get('username', '').strip()
    if user_filter:
        qs = qs.filter(username__icontains=user_filter)

    # Estatísticas do período
    all_period = AccessLog.objects.filter(timestamp__gte=date_from, timestamp__lte=date_to)
    stats = {
        'total':      all_period.count(),
        'login_ok':   all_period.filter(event=AccessLog.EVENT_LOGIN_OK).count(),
        'login_fail': all_period.filter(event=AccessLog.EVENT_LOGIN_FAIL).count(),
        'logout':     all_period.filter(event=AccessLog.EVENT_LOGOUT).count(),
        'lockout':    all_period.filter(event=AccessLog.EVENT_LOCKOUT).count(),
    }

    # Converter timestamps para SP antes de passar ao template
    logs = []
    for log in qs[:500]:  # limite de 500 para a listagem
        logs.append({
            'id':         log.id,
            'username':   log.username,
            'event':      log.event,
            'event_label': log.get_event_display(),
            'ip_address': log.ip_address or '—',
            'user_agent': log.user_agent[:80] + '…' if len(log.user_agent) > 80 else log.user_agent,
            'timestamp':  _format_timestamp(log.timestamp),
        })

    now_sp = dj_timezone.now().astimezone(SP_TZ)

    ctx = _settings_base_ctx()
    ctx.update({
        'active_tab':    'logs',
        'logs':          logs,
        'stats':         stats,
        'date_from_str': date_from.astimezone(SP_TZ).strftime('%Y-%m-%d'),
        'date_to_str':   date_to.astimezone(SP_TZ).strftime('%Y-%m-%d'),
        'event_filter':  event_filter,
        'user_filter':   user_filter,
        'period':        request.GET.get('period', ''),
        'now_sp':        now_sp.strftime('%d/%m/%Y %H:%M'),
        'event_choices': AccessLog.EVENT_CHOICES,
    })
    return render(request, 'settings.html', ctx)


# ── Exportação de Logs ────────────────────────────────────────────
def _get_export_qs(request):
    date_from, date_to = _parse_log_filters(request)
    qs = AccessLog.objects.filter(
        timestamp__gte=date_from,
        timestamp__lte=date_to,
    ).select_related('user').order_by('-timestamp')
    event_filter = request.GET.get('event', '')
    if event_filter:
        qs = qs.filter(event=event_filter)
    user_filter = request.GET.get('username', '').strip()
    if user_filter:
        qs = qs.filter(username__icontains=user_filter)
    return qs


EXPORT_HEADERS = ['Data/Hora (SP)', 'Usuário', 'Evento', 'IP', 'User Agent']


def _export_filename(ext):
    """Gera nome de arquivo: {sistema}_logs_{YYYYMMDD}.{ext}"""
    system_name = get_config('system_name', 'MyApp').strip() or 'MyApp'
    # Remove caracteres inválidos para nome de arquivo
    safe_name = ''.join(c if c.isalnum() or c in '-_' else '_' for c in system_name)
    date_str = dj_timezone.now().astimezone(SP_TZ).strftime('%Y%m%d')
    return f'{safe_name}_logs_{date_str}.{ext}'


def _log_rows(qs):
    for log in qs:
        yield [
            _format_timestamp(log.timestamp),
            log.username,
            log.get_event_display(),
            log.ip_address or '',
            log.user_agent,
        ]


@require_permission('settings')
def export_logs_csv(request):
    qs = _get_export_qs(request)
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{_export_filename("csv")}"'
    writer = csv.writer(response)
    writer.writerow(EXPORT_HEADERS)
    for row in _log_rows(qs):
        writer.writerow(row)
    return response


@require_permission('settings')
def export_logs_txt(request):
    qs = _get_export_qs(request)
    lines = ['\t'.join(EXPORT_HEADERS)]
    for row in _log_rows(qs):
        lines.append('\t'.join(str(c) for c in row))
    content = '\r\n'.join(lines)
    response = HttpResponse(content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{_export_filename("txt")}"'
    return response


@require_permission('settings')
def export_logs_xlsx(request):
    """Exporta XLSX usando apenas stdlib (sem openpyxl): gera CSV com .xlsx extension
    ou usa openpyxl se disponível."""
    qs = _get_export_qs(request)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Logs de Acesso'

        # Cabeçalho estilizado
        header_fill = PatternFill('solid', fgColor='6C63FF')
        header_font = Font(bold=True, color='FFFFFF')
        for col, header in enumerate(EXPORT_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(_log_rows(qs), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Larguras das colunas
        col_widths = [22, 18, 22, 16, 60]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{_export_filename("xlsx")}"'
        return response

    except ImportError:
        # Fallback: CSV com extensão xlsx se openpyxl não estiver instalado
        return export_logs_csv(request)
